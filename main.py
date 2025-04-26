import pandas as pd
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import os
import pyautogui

def connectvpn():
    original_window = driver.current_window_handle
    retries = 3

    for attempt in range(retries):
        try:
            pyautogui.hotkey('ctrl', 'q')
            driver.switch_to.window(driver.window_handles[-1])
            connect_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "ConnectionButton"))
            )
            connect_button.click()
            time.sleep(8)
            print("VPN connected successfully")
            break
        except Exception as e:
            print(f"Attempt {attempt + 1} failed to connect VPN: {e}")
            if attempt == retries - 1:
                print("Max retries reached. Failed to connect VPN.")
        finally:
            driver.switch_to.window(original_window)

def discountvpn():
    original_window = driver.current_window_handle
    retries = 3

    for attempt in range(retries):
        try:
            pyautogui.hotkey('ctrl', 'q')
            driver.switch_to.window(driver.window_handles[-1])
            disconnect_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="ConnectionButton"]/div[3]'))
            )
            disconnect_button.click()
            time.sleep(5)
            print("VPN disconnected successfully")
            break
        except Exception as e:
            print(f"Attempt {attempt + 1} failed to disconnect VPN: {e}")
            if attempt == retries - 1:
                print("Max retries reached. Failed to disconnect VPN.")
        finally:
            driver.switch_to.window(original_window)

def check_xpath_exists(driver, xpath):
    try:
        element = driver.find_element(By.XPATH, xpath)
        return element.is_displayed()
    except Exception:
        return False

def check_css_selector_data(driver, selector):
    try:
        element = driver.find_element(By.CSS_SELECTOR, selector)
        return element.is_displayed()
    except Exception:
        return False


# Initialize and load Excel data
excel_file = "test.xlsx"
data = pd.read_excel(excel_file)
wb = load_workbook(excel_file)
ws = wb.active

# Setup Chrome with VPN extension
options = webdriver.ChromeOptions()
touchvpn_path = 'touchvpn.crx'
if not os.path.exists(touchvpn_path):
    raise FileNotFoundError(f"TouchVPN CRX file not found at: {touchvpn_path}")
options.add_extension(touchvpn_path)

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

try:
    driver.get("chrome://extensions/shortcuts")
    time.sleep(4)
    pyautogui.hotkey('ctrl', 'w')
    time.sleep(1)

    for _ in range(5):
        pyautogui.press('tab')
        time.sleep(0.2)
    pyautogui.press('enter')
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'q')
    time.sleep(0.2)
    pyautogui.press('enter')

    driver.execute_script("window.open('https://enquiry.navigate.mib.org.uk/checkyourvehicle','_blank');")
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(2)

    #connectvpn()
    driver.refresh()
    time.sleep(2)

    for index, row in data.iterrows():
        input_value = row['Registration']
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(2)

        try:
            input_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div[1]/div/div[2]/div[2]/div[3]/div/input'))
            )
            input_element.clear()
            input_element.send_keys(input_value)

            actions = ActionChains(driver)
            actions.scroll_by_amount(0, 1000).perform()
            time.sleep(1.5)

            checkbox = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/main/div[1]/div/div[2]/div[2]/label/input'))
            )
            checkbox.click()
            time.sleep(0.5)

            submit_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/main/div[1]/div/div[2]/div[2]/div[7]/button'))
            )
            submit_button.click()
            time.sleep(5)

            xpath_to_check = '//*[@id="__next"]/main/div[1]/div/div[2]/div[2]/div[3]/div/div[2]/span[2]'
            selector_data = "#__next > main > div.container > div > div:nth-child(2) > div.p-md-4.py-2.col-md-8 > div > div > svg > path:nth-child(2)"

            while check_xpath_exists(driver, xpath_to_check):
                discountvpn()
                time.sleep(1)
                connectvpn()
                driver.refresh()
                time.sleep(2)

                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(2)

                input_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div[1]/div/div[2]/div[2]/div[3]/div/input'))
                )
                input_element.clear()
                input_element.send_keys(input_value)

                actions = ActionChains(driver)
                actions.scroll_by_amount(0, 1000).perform()
                time.sleep(1.5)

                checkbox = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/main/div[1]/div/div[2]/div[2]/label/input'))
                )
                checkbox.click()
                time.sleep(0.5)

                submit_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/main/div[1]/div/div[2]/div[2]/div[7]/button'))
                )
                submit_button.click()
                time.sleep(5)

                if not check_xpath_exists(driver, xpath_to_check):
                    break

            if check_css_selector_data(driver, selector_data):
                result = "Found"
                print("Found -> " + input_value)
            else:
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(1.5)
                result = "Not Found"
                print("Not Found -> " + input_value)

            ws[f"C{index + 2}"] = result
            wb.save(excel_file)
            driver.refresh()
            time.sleep(2)

        except Exception as e:
            print(f"Error processing registration {input_value}: {e}")
            ws[f"C{index + 2}"] = "Error"
            wb.save(excel_file)

finally:
    time.sleep(50)
    driver.quit()